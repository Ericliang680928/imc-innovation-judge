"""
批次評審處理器
─────────────────────────────────────
- 同時處理多份檔案
- 產出每份 .docx + .pdf
- 額外產出一份 .xlsx 彙整表(各件分數對照)
- 整體打包成 .zip 下載
"""

import io
import time
import zipfile
from datetime import datetime
from file_readers import read_file
from report_generator import markdown_to_docx, markdown_to_pdf


def build_batch_zip(
    batch_results: list,
) -> bytes:
    """
    輸入:每件評審結果 list,每筆為:
      {
        "file_name": str,
        "case_title": str,
        "category": str,
        "stage": str,
        "report_markdown": str,
        "panel_score": dict (optional, 多評審模式才有),
        "per_judge_data": list (optional, 多評審),
      }
    回傳:ZIP bytes,包含:
      - 每件的 .md / .docx / .pdf
      - 一份「批次評審彙整.xlsx」
      - 一份「批次評審彙整.md」
    """
    buf = io.BytesIO()

    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for i, r in enumerate(batch_results, 1):
            base = _safe_filename(f"{i:02d}_{r['file_name']}")
            meta = {
                "category": r["category"],
                "stage": r["stage"],
                "case_title": r.get("case_title", ""),
                "file_name": r["file_name"],
            }

            # Markdown
            zf.writestr(f"{base}/report.md", r["report_markdown"].encode("utf-8"))

            # Word
            try:
                docx_bytes = markdown_to_docx(r["report_markdown"], meta)
                zf.writestr(f"{base}/report.docx", docx_bytes)
            except Exception as e:
                zf.writestr(f"{base}/report.docx.ERROR.txt",
                            f"Word 產生失敗:{e}".encode("utf-8"))

            # PDF
            try:
                pdf_bytes = markdown_to_pdf(r["report_markdown"], meta)
                zf.writestr(f"{base}/report.pdf", pdf_bytes)
            except Exception as e:
                zf.writestr(f"{base}/report.pdf.ERROR.txt",
                            f"PDF 產生失敗:{e}".encode("utf-8"))

        # 彙整 Excel
        try:
            xlsx_bytes = _build_summary_xlsx(batch_results)
            zf.writestr("00_批次評審彙整.xlsx", xlsx_bytes)
        except Exception as e:
            zf.writestr("00_批次評審彙整.xlsx.ERROR.txt",
                        f"Excel 產生失敗:{e}".encode("utf-8"))

        # 彙整 Markdown
        summary_md = _build_summary_markdown(batch_results)
        zf.writestr("00_批次評審彙整.md", summary_md.encode("utf-8"))

    return buf.getvalue()


def _safe_filename(name: str) -> str:
    """移除 Windows/Mac/Linux 不允許的字元。"""
    import re
    # 移除副檔名
    name = re.sub(r"\.[a-zA-Z0-9]{1,5}$", "", name)
    # 移除非法字元
    name = re.sub(r'[\\/:*?"<>|]+', "_", name)
    # 避免太長
    return name[:80].strip()


# ─────────────────────────────────────────────
# Excel 彙整表
# ─────────────────────────────────────────────
def _build_summary_xlsx(batch_results: list) -> bytes:
    """產生批次評審彙整 Excel。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "批次評審彙整"

    # 標題列
    headers = ["編號", "檔案名稱", "案件題目", "類組", "階段", "總分", "滿分", "達成率"]
    # 是否為多評審模式(找第一筆有 per_judge_data 的)
    has_panel = any(r.get("panel_score") for r in batch_results)
    if has_panel:
        headers.extend(["評審數", "最高分(去除)", "最低分(去除)", "原始平均"])

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF", size=12)
        cell.fill = PatternFill("solid", fgColor="2A6CA8")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 資料列
    for row_i, r in enumerate(batch_results, 2):
        panel = r.get("panel_score") or {}
        total = panel.get("trimmed_average") or panel.get("raw_average") or r.get("total")
        max_total = panel.get("max_total") or r.get("max_total")
        rate = f"{round(total/max_total*100, 1)}%" if (total and max_total) else ""

        row = [
            row_i - 1,
            r["file_name"],
            r.get("case_title", "") or "(未填)",
            r["category"],
            r["stage"],
            total,
            max_total,
            rate,
        ]
        if has_panel:
            row.extend([
                len(panel.get("totals", [])) if panel else 1,
                f"{panel['max'][1]} ({panel['max'][0]})" if panel.get("max") else "",
                f"{panel['min'][1]} ({panel['min'][0]})" if panel.get("min") else "",
                panel.get("raw_average", ""),
            ])

        for col, val in enumerate(row, 1):
            cell = ws.cell(row=row_i, column=col, value=val)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if col == 6:  # 總分欄加粗
                cell.font = Font(bold=True, size=12, color="1A4D7A")

    # 欄寬調整
    widths = [6, 32, 28, 20, 16, 8, 8, 10, 8, 18, 18, 10]
    for i, w in enumerate(widths[:len(headers)], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 28

    # 凍結首列
    ws.freeze_panes = "A2"

    # 儲存
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_summary_markdown(batch_results: list) -> str:
    """產生批次評審彙整 Markdown。"""
    lines = [
        f"# 🏆 IMC 第 21 屆創新獎 — 批次評審彙整",
        "",
        f"- **批次評審件數**:{len(batch_results)} 件",
        f"- **產生時間**:{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        "",
        "---",
        "",
        "## 📊 各案件總分對照",
        "",
        "| 編號 | 檔案名稱 | 類組 | 階段 | 總分 | 滿分 | 達成率 |",
        "|---|---|---|---|---|---|---|",
    ]

    for i, r in enumerate(batch_results, 1):
        panel = r.get("panel_score") or {}
        total = panel.get("trimmed_average") or panel.get("raw_average") or r.get("total")
        max_total = panel.get("max_total") or r.get("max_total")
        rate = f"{round(total/max_total*100, 1)}%" if (total and max_total) else "—"
        lines.append(
            f"| {i} | {r['file_name']} | {r['category']} | {r['stage']} | "
            f"**{total}** | {max_total} | {rate} |"
        )
    lines.append("")

    # 排名
    ranked = sorted(
        batch_results,
        key=lambda x: -(
            (x.get("panel_score") or {}).get("trimmed_average")
            or (x.get("panel_score") or {}).get("raw_average")
            or x.get("total") or 0
        )
    )
    lines.extend([
        "## 🏅 名次排序",
        "",
    ])
    for rank, r in enumerate(ranked, 1):
        panel = r.get("panel_score") or {}
        total = panel.get("trimmed_average") or panel.get("raw_average") or r.get("total")
        medal = {1: "🥇", 2: "🥈", 3: "🥉"}.get(rank, f"#{rank}")
        lines.append(f"- {medal} **{total} 分** | {r['file_name']} | {r.get('case_title', '') or '(未填)'}")
    lines.append("")

    return "\n".join(lines)
